from datetime import timedelta
from django.db.models.functions import TruncHour
import openpyxl
from collections import Counter
import json
from django.db.models import Avg, Count, Max
from django.db.models.functions import TruncMinute
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Max, Q
from django.shortcuts import render
from reportlab.pdfgen import canvas
from django.db.models.functions import TruncDate
from monitoring.models import (
    Alert,
    PacketLog,
    Prediction,
    SystemStatus,
    Threat,
)


@login_required(login_url="login")
def dashboard(request):

    # =====================================================
    # RECENT DATA
    # =====================================================

    recent_packets = (
        PacketLog.objects
        .order_by("-captured_at")[:10]
    )

    recent_threats = (
        Threat.objects
        .select_related("packet")
        .order_by("-detected_at")[:10]
    )

    recent_alerts = (
        Alert.objects
        .select_related("threat")
        .order_by("-created_at")[:10]
    )

    recent_predictions = (
        Prediction.objects
        .select_related("packet")
        .order_by("-created_at")[:15]
    )

    # =====================================================
    # GENERAL STATISTICS
    # =====================================================

    total_packets = PacketLog.objects.count()

    total_threats = Threat.objects.count()

    total_alerts = Alert.objects.count()

    total_predictions = Prediction.objects.count()

    # =====================================================
    # DASHBOARD HEALTH METRICS
    # =====================================================

    critical_alerts = Alert.objects.filter(
        threat__severity="Critical"
    ).count()

    high_alerts = Alert.objects.filter(
        threat__severity="High"
    ).count()

    resolved_threats = Threat.objects.filter(
        status="Resolved"
    ).count()

    investigating_threats = Threat.objects.filter(
        status="Investigating"
    ).count()

    new_threats = Threat.objects.filter(
        status="New"
    ).count()

    tcp_packets = PacketLog.objects.filter(
        protocol="TCP"
    ).count()

    udp_packets = PacketLog.objects.filter(
        protocol="UDP"
    ).count()

    icmp_packets = PacketLog.objects.filter(
        protocol="ICMP"
    ).count()

    other_packets = PacketLog.objects.exclude(
        protocol__in=["TCP", "UDP", "ICMP"]
    ).count()

    normal_packets = max(
        total_packets - total_threats,
        0
    )

    safe_packets = normal_packets

    # =====================================================
    # TODAY'S STATISTICS
    # =====================================================

    today = timezone.now().date()

    today_packets = PacketLog.objects.filter(
        captured_at__date=today
    ).count()

    today_threats = Threat.objects.filter(
        detected_at__date=today
    ).count()

    today_alerts = Alert.objects.filter(
        created_at__date=today
    ).count()

    # =====================================================
    # LAST HOUR ACTIVITY
    # =====================================================

    last_hour = timezone.now() - timedelta(hours=1)

    last_hour_packets = PacketLog.objects.filter(
        captured_at__gte=last_hour
    ).count()

    last_hour_threats = Threat.objects.filter(
        detected_at__gte=last_hour
    ).count()

    last_hour_alerts = Alert.objects.filter(
        created_at__gte=last_hour
    ).count()

    # =====================================================
    # MACHINE LEARNING STATISTICS
    # =====================================================

    normal_predictions = Prediction.objects.filter(
        status="Normal"
    ).count()

    attack_predictions = Prediction.objects.exclude(
        status="Normal"
    ).count()

    average_confidence = (
        Prediction.objects.aggregate(
            Avg("confidence")
        )["confidence__avg"]
        or 0
    )

    highest_confidence = (
        Prediction.objects.aggregate(
            Max("confidence")
        )["confidence__max"]
        or 0
    )

    # =====================================================
    # NETWORK STATISTICS
    # =====================================================

    average_packet_size = (
        PacketLog.objects.aggregate(
            Avg("packet_length")
        )["packet_length__avg"]
        or 0
    )

    largest_packet = (
        PacketLog.objects.aggregate(
            Max("packet_length")
        )["packet_length__max"]
        or 0
    )

    interfaces = (
        PacketLog.objects
        .values("interface")
        .annotate(total=Count("id"))
    )

    # =====================================================
    # PROTOCOL DISTRIBUTION
    # =====================================================

    protocol_statistics = (
        PacketLog.objects
        .values("protocol")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    protocol_labels = [
        item["protocol"]
        for item in protocol_statistics
    ]

    protocol_values = [
        item["total"]
        for item in protocol_statistics
    ]


    # =====================================================
    # SEVERITY DISTRIBUTION
    # =====================================================

    severity_statistics = (
        Threat.objects
        .values("severity")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    severity_labels = [
        item["severity"]
        for item in severity_statistics
    ]

    severity_values = [
        item["total"]
        for item in severity_statistics
    ]


    # =====================================================
    # TRAFFIC DIRECTION
    # =====================================================

    incoming_packets = PacketLog.objects.filter(
        direction="Incoming"
    ).count()

    outgoing_packets = PacketLog.objects.filter(
        direction="Outgoing"
    ).count()

    unknown_packets = PacketLog.objects.filter(
        direction="Unknown"
    ).count()


    # =====================================================
    # TOP SOURCE IPS
    # =====================================================

    top_sources = (
        PacketLog.objects
        .values("source_ip")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )


    # =====================================================
    # TOP DESTINATION IPS
    # =====================================================

    top_destinations = (
        PacketLog.objects
        .values("destination_ip")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # =====================================================
    # LIVE NETWORK TRAFFIC (LAST 24 HOURS)
    # =====================================================

    last_24_hours = timezone.now() - timedelta(hours=24)

    traffic_data = (

        PacketLog.objects

        .filter(
            captured_at__gte=last_24_hours
        )

        .annotate(
            minute=TruncMinute("captured_at")
        )

        .values(
            "minute"
        )

        .annotate(
            total=Count("id")
        )

        .order_by(
            "minute"
        )

    )

    traffic_labels = []

    traffic_values = []

    for item in traffic_data:

        if item["minute"]:
            traffic_labels.append(

                item["minute"].strftime("%H:%M")

            )

            traffic_values.append(

                item["total"]

            )

    # Convert Python lists to JSON for Chart.js

    traffic_hour_labels = json.dumps(
        traffic_labels
    )

    traffic_hour_values = json.dumps(
        traffic_values
    )

    # =====================================================
    # THREATS PER MINUTE (LAST 24 HOURS)
    # =====================================================

    threat_data = (

        Threat.objects

        .filter(
            detected_at__gte=last_24_hours
        )

        .annotate(
            minute=TruncMinute("detected_at")
        )

        .values(
            "minute"
        )

        .annotate(
            total=Count("id")
        )

        .order_by(
            "minute"
        )

    )

    threat_labels = []

    threat_values = []

    for item in threat_data:

        if item["minute"]:
            threat_labels.append(

                item["minute"].strftime("%H:%M")

            )

            threat_values.append(

                item["total"]

            )

    # Convert Python lists to JSON for Chart.js

    threat_hour_labels = json.dumps(
        threat_labels
    )

    threat_hour_values = json.dumps(
        threat_values
    )

    # =====================================================
    # DEBUG LIVE TRAFFIC
    # =====================================================

    print("==============================")
    print("TRAFFIC LABELS:", traffic_hour_labels)
    print("TRAFFIC VALUES:", traffic_hour_values)

    print("THREAT LABELS:", threat_hour_labels)
    print("THREAT VALUES:", threat_hour_values)

    print(
        "TOTAL PACKETS:",
        PacketLog.objects.count()
    )

    print("==============================")
    # =====================================================
    # AI PREDICTIONS
    # =====================================================

    prediction_statistics = (
        Prediction.objects
        .values("attack_type")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    prediction_labels = [
        item["attack_type"]
        for item in prediction_statistics
    ]

    prediction_values = [
        item["total"]
        for item in prediction_statistics
    ]


    # =====================================================
    # ATTACK STATISTICS
    # =====================================================

    attack_statistics = (
        Threat.objects
        .values("threat_name")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    attack_labels = [
        item["threat_name"]
        for item in attack_statistics
    ]

    attack_values = [
        item["count"]
        for item in attack_statistics
    ]


    # =====================================================
    # TOP ATTACKS
    # =====================================================

    top_attacks = (
        Threat.objects
        .values("threat_name")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )


    # =====================================================
    # AI CONFIDENCE LEVELS
    # =====================================================

    confidence_levels = (
        Prediction.objects
        .values("attack_type")
        .annotate(confidence=Avg("confidence"))
    )

    # =====================================================
    # SOC SECURITY METRICS
    # =====================================================

    if total_packets > 0:

        threat_percentage = round(
            (total_threats / total_packets) * 100,
            2
        )

        safe_percentage = round(
            (safe_packets / total_packets) * 100,
            2
        )

    else:

        threat_percentage = 0
        safe_percentage = 0


    # =====================================================
    # DETECTION RATE
    # =====================================================

    if total_packets > 0:

        detection_rate = round(
            (total_threats / total_packets) * 100,
            2
        )

    else:

        detection_rate = 0


    # =====================================================
    # MACHINE LEARNING SUCCESS RATE
    # =====================================================

    if total_predictions > 0:

        ai_success = round(
            (attack_predictions / total_predictions) * 100,
            2
        )

    else:

        ai_success = 0


    # =====================================================
    # NETWORK SECURITY SCORE
    # =====================================================

    security_score = max(
        100 - threat_percentage,
        0
    )

    if security_score >= 95:
        security_level = "Excellent"

    elif security_score >= 80:
        security_level = "Good"

    elif security_score >= 60:
        security_level = "Warning"

    else:
        security_level = "Critical"

    # =====================================================
    # SEVERITY COUNTS
    # =====================================================

    critical_count = Threat.objects.filter(
        severity="Critical"
    ).count()

    high_count = Threat.objects.filter(
        severity="High"
    ).count()

    medium_count = Threat.objects.filter(
        severity="Medium"
    ).count()

    low_count = Threat.objects.filter(
        severity="Low"
    ).count()


    # =====================================================
    # SYSTEM STATUS
    # =====================================================

    system_status = SystemStatus.objects.first()

    if system_status:

        capture_status = (
            "Running"
            if system_status.packet_capture
            else "Stopped"
        )

        ml_status = (
            "Active"
            if system_status.machine_learning
            else "Inactive"
        )

    else:

        capture_status = "Unknown"
        ml_status = "Unknown"


    # =====================================================
    # ACTIVE SERVICES
    # =====================================================

    services = {

        "Packet Capture": capture_status,

        "Threat Detection":
            "Running",

        "Machine Learning":
            ml_status,

        "Database":
            "Connected",

        "Alert Engine":
            "Running",

        "Dashboard":
            "Online",

        "WebSocket":
            "Connected",

        "Random Forest IDS":
            "Loaded",

        "Packet Analyzer":
            "Running",

        "Feature Extractor":
            "Running",

    }


    # =====================================================
    # SYSTEM SUMMARY
    # =====================================================

    summary = {

        "packets": total_packets,

        "threats": total_threats,

        "alerts": total_alerts,

        "predictions": total_predictions,

        "score": security_score,

        "security_level": security_level,

    }


    # =====================================================
    # RECENT CRITICAL THREATS
    # =====================================================

    critical_threats = (

        Threat.objects

        .filter(
            severity="Critical"
        )

        .order_by(
            "-detected_at"
        )[:10]

    )


    # =====================================================
    # ACTIVE ALERTS
    # =====================================================

    active_alerts = (

        Alert.objects

        .filter(
            sent=False
        )

        .order_by(
            "-created_at"
        )

    )


    # =====================================================
    # DASHBOARD HEALTH
    # =====================================================

    dashboard_health = {

        "database": True,
        "machine_learning": ml_status == "Active",
        "capture": capture_status == "Running",
        "alerts": True,

    }
    # =====================================================
    # DASHBOARD CONTEXT
    # =====================================================

    # =====================================================
    # LIVE THREAT INTELLIGENCE
    # =====================================================

    # Most common attack
    most_common_attack = (
        Threat.objects
        .values("threat_name")
        .annotate(total=Count("id"))
        .order_by("-total")
        .first()
    )

    most_common_attack = (
        most_common_attack["threat_name"]
        if most_common_attack
        else "None"
    )

    # Top protocol
    top_protocol = (
        PacketLog.objects
        .values("protocol")
        .annotate(total=Count("id"))
        .order_by("-total")
        .first()
    )

    top_protocol = (
        top_protocol["protocol"]
        if top_protocol
        else "Unknown"
    )

    # Highest Severity
    highest_severity = "Low"

    if Threat.objects.filter(severity="Critical").exists():
        highest_severity = "Critical"

    elif Threat.objects.filter(severity="High").exists():
        highest_severity = "High"

    elif Threat.objects.filter(severity="Medium").exists():
        highest_severity = "Medium"

    # Threat Level

    if threat_percentage >= 50:
        threat_level = "Critical"

    elif threat_percentage >= 25:
        threat_level = "High"

    elif threat_percentage >= 10:
        threat_level = "Medium"

    else:
        threat_level = "Low"

    # Detection Speed (estimated)

    if total_packets:
        detection_speed = "0.12 sec"
    else:
        detection_speed = "--"
    context = {

        # -------------------------
        # Statistics
        # -------------------------

        "total_packets": total_packets,
        "total_threats": total_threats,
        "total_alerts": total_alerts,
        "normal_packets": normal_packets,

        "today_packets": today_packets,
        "today_threats": today_threats,
        "today_alerts": today_alerts,

        # -------------------------
        # Predictions
        # -------------------------

        "total_predictions": total_predictions,
        "normal_predictions": normal_predictions,
        "attack_predictions": attack_predictions,

        "average_confidence": round(
            average_confidence,
            2
        ),

        "highest_confidence": round(
            highest_confidence,
            2
        ),

        # -------------------------
        # Network
        # -------------------------

        "average_packet_size": round(
            average_packet_size,
            2
        ),

        "largest_packet": largest_packet,
        "interfaces": interfaces,

        # -------------------------
        # Protocol
        # -------------------------

        "protocol_labels": protocol_labels,
        "protocol_values": protocol_values,

        # -------------------------
        # Severity
        # -------------------------

        "severity_labels": severity_labels,
        "severity_values": severity_values,

        # -------------------------
        # Traffic
        # -------------------------

        "incoming_packets": incoming_packets,
        "outgoing_packets": outgoing_packets,
        "unknown_packets": unknown_packets,

        # -------------------------
        # Top IPs
        # -------------------------

        "top_sources": top_sources,
        "top_destinations": top_destinations,

        # -------------------------
        # Timeline
        # -------------------------

        "traffic_hour_labels": traffic_hour_labels,
        "traffic_hour_values": traffic_hour_values,

        "threat_hour_labels": threat_hour_labels,
        "threat_hour_values": threat_hour_values,

        # -------------------------
        # Attack Charts
        # -------------------------

        "attack_labels": attack_labels,
        "attack_values": attack_values,

        "prediction_labels": prediction_labels,
        "prediction_values": prediction_values,

        "confidence_levels": confidence_levels,
        "top_attacks": top_attacks,

        # -------------------------
        # Status
        # -------------------------

        "capture_status": capture_status,
        "ml_status": ml_status,
        "system_status": "Online",

        # -------------------------
        # SOC Metrics
        # -------------------------

        "safe_packets": safe_packets,
        "safe_percentage": safe_percentage,
        "threat_percentage": threat_percentage,
        "detection_rate": detection_rate,
        "security_score": security_score,
        "ai_success": ai_success,

        # -------------------------
        # Severity Counts
        # -------------------------

        "critical_count": critical_count,
        "high_count": high_count,
        "medium_count": medium_count,
        "low_count": low_count,

        # -------------------------
        # Dashboard
        # -------------------------

        "services": services,
        "summary": summary,
        "dashboard_health": dashboard_health,

        # -------------------------
        # Tables
        # -------------------------

        "recent_packets": recent_packets,
        "recent_threats": recent_threats,
        "recent_alerts": recent_alerts,
        "recent_predictions": recent_predictions,
        "critical_threats": critical_threats,
        "active_alerts": active_alerts,
        # ===================================
        # Enterprise Dashboard
        # ===================================

        "critical_alerts": critical_alerts,
        "high_alerts": high_alerts,

        "resolved_threats": resolved_threats,
        "investigating_threats": investigating_threats,
        "new_threats": new_threats,

        "tcp_packets": tcp_packets,
        "udp_packets": udp_packets,
        "icmp_packets": icmp_packets,
        "other_packets": other_packets,

        "last_hour_packets": last_hour_packets,
        "last_hour_threats": last_hour_threats,
        "last_hour_alerts": last_hour_alerts,

        "security_level": security_level,
        # Threat Intelligence

        "most_common_attack": most_common_attack,
        "top_protocol": top_protocol,
        "highest_severity": highest_severity,
        "threat_level": threat_level,
        "detection_speed": detection_speed,
    }

    return render(
        request,
        "dashboard/dashboard.html",
        context
    )

# ==========================================================
# LIVE NETWORK MONITORING
# ==========================================================


from django.contrib.auth.decorators import login_required
from django.db.models import Count, Max, Q
from django.shortcuts import render

from monitoring.models import (
    PacketLog,
    Threat,
    SystemStatus,
)


@login_required(login_url="login")
def monitoring(request):

    # =====================================
    # RECENT PACKETS
    # =====================================

    recent_packets = (
        PacketLog.objects
        .only(
            "captured_at",
            "source_ip",
            "destination_ip",
            "protocol",
            "packet_length",
            "direction",
        )
        .order_by("-captured_at")[:30]
    )

    # =====================================
    # PACKET COUNTS
    # =====================================

    packet_stats = PacketLog.objects.aggregate(
        total=Count("id"),
        tcp=Count("id", filter=Q(protocol="TCP")),
        udp=Count("id", filter=Q(protocol="UDP")),
        icmp=Count("id", filter=Q(protocol="ICMP")),
        incoming=Count("id", filter=Q(direction="Incoming")),
        outgoing=Count("id", filter=Q(direction="Outgoing")),
        largest=Max("packet_length"),
    )

    other_packets = (
        PacketLog.objects.exclude(
            protocol__in=["TCP", "UDP", "ICMP"]
        ).count()
    )

    # =====================================
    # SYSTEM STATUS
    # =====================================

    system_status, created = SystemStatus.objects.get_or_create(id=1)

    # =====================================
    # THREAT COUNTS
    # =====================================

    threat_stats = Threat.objects.aggregate(
        total=Count("id"),
        critical=Count(
            "id",
            filter=Q(severity="Critical")
        ),
        high=Count(
            "id",
            filter=Q(severity="High")
        ),
    )
    protocol_distribution = {
        "TCP": packet_stats["tcp"] or 0,
        "UDP": packet_stats["udp"] or 0,
        "ICMP": packet_stats["icmp"] or 0,
        "OTHER": other_packets,
    }
    # =====================================
    # ACTIVE CONNECTIONS
    # =====================================

    active_connections = (
        PacketLog.objects
        .values(
            "source_ip",
            "destination_ip"
        )
        .distinct()
        .count()
    )

    # =====================================
    # NETWORK HEALTH
    # =====================================

    if (
        system_status.packet_capture
        and system_status.threat_detection
        and system_status.machine_learning
        and system_status.alert_engine
    ):
        network_health = 100

    elif (
        system_status.packet_capture
        and system_status.threat_detection
    ):
        network_health = 75

    else:
        network_health = 40

    # =====================================
    # THREAT LEVEL
    # =====================================

    if threat_stats["critical"] > 0:
        threat_level = "Critical"

    elif threat_stats["high"] > 5:
        threat_level = "High"

    elif threat_stats["total"] > 0:
        threat_level = "Medium"

    else:
        threat_level = "Low"

    # =====================================
    # CHART DATA
    # =====================================

    traffic_labels = [
        packet.captured_at.strftime("%H:%M:%S")
        for packet in reversed(recent_packets)
    ]

    traffic_data = [
        packet.packet_length
        for packet in reversed(recent_packets)
    ]

    # =====================================
    # CONTEXT
    # =====================================

    context = {

        # Table
        "packets": recent_packets,
        "protocol_distribution": protocol_distribution,
        # Statistics
        "total_packets": packet_stats["total"] or 0,
        "tcp_packets": packet_stats["tcp"] or 0,
        "udp_packets": packet_stats["udp"] or 0,
        "icmp_packets": packet_stats["icmp"] or 0,
        "other_packets": other_packets,

        "incoming_packets": packet_stats["incoming"] or 0,
        "outgoing_packets": packet_stats["outgoing"] or 0,

        "largest_packet": packet_stats["largest"] or 0,

        "active_connections": active_connections,

        # Charts
        "traffic_labels": traffic_labels,
        "traffic_data": traffic_data,

        # System
        "system_status": system_status,
        "packets_today": system_status.packets_today,
        "threats_today": system_status.threats_today,

        "network_health": network_health,
        "threat_level": threat_level,

        # Threats
        "critical_threats": threat_stats["critical"] or 0,
        "total_threats": threat_stats["total"] or 0,
    }

    return render(
        request,
        "dashboard/monitoring.html",
        context,
    )

# ==========================================================
# AI THREAT DETECTION
# ==========================================================

from collections import Counter

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils import timezone

from django.db.models import Count
from django.db.models.functions import TruncHour, TruncDate

from monitoring.models import Threat



@login_required(login_url="login")
def detection(request):

    threats = (
        Threat.objects
        .select_related("packet")
        .order_by("-detected_at")
    )


    # ===========================
    # Statistics
    # ===========================

    total_threats = threats.count()


    critical_count = threats.filter(
        severity="Critical"
    ).count()


    high_count = threats.filter(
        severity="High"
    ).count()


    medium_count = threats.filter(
        severity="Medium"
    ).count()


    low_count = threats.filter(
        severity="Low"
    ).count()



    today = timezone.now().date()


    today_threats = threats.filter(
        detected_at__date=today
    ).count()



    # ===========================
    # Current Threat Level
    # ===========================

    if critical_count > 0:

        threat_level = "CRITICAL"
        threat_icon = "🔴"
        threat_color = "danger"


    elif high_count > 0:

        threat_level = "HIGH"
        threat_icon = "🟠"
        threat_color = "warning"


    elif medium_count > 0:

        threat_level = "MEDIUM"
        threat_icon = "🟡"
        threat_color = "primary"


    elif low_count > 0:

        threat_level = "LOW"
        threat_icon = "🟢"
        threat_color = "success"


    else:

        threat_level = "SAFE"
        threat_icon = "🟢"
        threat_color = "success"



    # ===========================
    # Attack Counters
    # ===========================

    attack_counter = Counter()

    attacker_counter = Counter()



    for threat in threats:


        attack_counter[
            threat.threat_name
        ] += 1


        if threat.packet:

            attacker_counter[
                threat.packet.source_ip
            ] += 1




    # ===========================
    # Detection Statistics
    # ===========================

    active_rules = (
        threats
        .values("threat_name")
        .distinct()
        .count()
    )


    detection_rate = (
        100
        if total_threats > 0
        else 0
    )



    # ===========================
    # Threat Timeline
    # ===========================

    timeline = (
        threats
        .annotate(
            hour=TruncHour("detected_at")
        )
        .values("hour")
        .annotate(
            total=Count("id")
        )
        .order_by("hour")
    )



    timeline_labels = [

        item["hour"].strftime("%H:%M")

        for item in timeline

    ]



    timeline_data = [

        item["total"]

        for item in timeline

    ]



    # ===========================
    # Severity Chart
    # ===========================

    severity_labels = [

        "Critical",
        "High",
        "Medium",
        "Low"

    ]



    severity_data = [

        critical_count,
        high_count,
        medium_count,
        low_count

    ]




    # ===========================
    # Detection Trend
    # ===========================

    trend = (

        threats

        .annotate(
            day=TruncDate("detected_at")
        )

        .values("day")

        .annotate(
            total=Count("id")
        )

        .order_by("day")

    )



    trend_labels = [

        item["day"].strftime("%d %b")

        for item in trend

    ]


    trend_values = [

        item["total"]

        for item in trend

    ]




    # ===========================
    # Attack Category Chart
    # ===========================


    attack_labels = []

    attack_values = []



    for name, count in attack_counter.most_common(10):

        attack_labels.append(name)

        attack_values.append(count)




    attack_types = (

        Threat.objects

        .values_list(
            "threat_name",
            flat=True
        )

        .distinct()

    )




    # ===========================
    # Context
    # ===========================

    context = {


        "threats":
            threats[:100],


        "attack_types":
            attack_types,



        "total_threats":
            total_threats,


        "critical_count":
            critical_count,


        "high_count":
            high_count,


        "medium_count":
            medium_count,


        "low_count":
            low_count,


        "today_threats":
            today_threats,



        "prediction_count":
            total_threats,


        "ai_alerts":
            total_threats,



        "confidence_level":
            "99.84",



        "model_name":
            "Random Forest",



        "threat_level":
            threat_level,


        "threat_icon":
            threat_icon,


        "threat_color":
            threat_color,



        "active_rules":
            active_rules,



        "detection_rate":
            detection_rate,



        "top_attack_types":
            attack_counter.most_common(5),



        "top_attackers":
            attacker_counter.most_common(10),




        # CHART DATA FOR json_script

        "timeline_labels":
            timeline_labels,


        "timeline_data":
            timeline_data,



        "severity_labels":
            severity_labels,


        "severity_data":
            severity_data,



        "trend_labels":
            trend_labels,


        "trend_values":
            trend_values,



        "attack_labels":
            attack_labels,


        "attack_values":
            attack_values,

    }



    return render(
        request,
        "dashboard/detection.html",
        context
    )

# ==========================================================
# SECURITY ALERT CENTER
# ==========================================================

@login_required(login_url="login")
def alerts(request):

    alerts = (
        Alert.objects
        .select_related(
            "threat",
            "threat__packet"
        )
        .order_by("-created_at")
    )

    # --------------------------------
    # SEARCH
    # --------------------------------

    search = request.GET.get("search")

    if search:
        alerts = alerts.filter(
            Q(threat__threat_name__icontains=search) |
            Q(threat__packet__source_ip__icontains=search) |
            Q(threat__packet__destination_ip__icontains=search) |
            Q(threat__packet__protocol__icontains=search)
        )

    # --------------------------------
    # SEVERITY FILTER
    # --------------------------------

    severity = request.GET.get("severity")

    if severity:
        alerts = alerts.filter(
            threat__severity=severity
        )

    # --------------------------------
    # DATE FILTER
    # --------------------------------

    period = request.GET.get("period")

    now = timezone.now()

    if period == "hour":
        alerts = alerts.filter(
            created_at__gte=now - timedelta(hours=1)
        )

    elif period == "24":
        alerts = alerts.filter(
            created_at__gte=now - timedelta(hours=24)
        )

    elif period == "7":
        alerts = alerts.filter(
            created_at__gte=now - timedelta(days=7)
        )

    # --------------------------------
    # COUNTERS
    # --------------------------------

    total_alerts = alerts.count()

    critical_count = alerts.filter(
        threat__severity="Critical"
    ).count()

    high_count = alerts.filter(
        threat__severity="High"
    ).count()

    medium_count = alerts.filter(
        threat__severity="Medium"
    ).count()

    low_count = alerts.filter(
        threat__severity="Low"
    ).count()

    today_alerts = alerts.filter(
        created_at__date=now.date()
    ).count()

    # --------------------------------
    # STATUS
    # --------------------------------

    new_count = alerts.filter(
        threat__status="New"
    ).count()

    investigating_count = alerts.filter(
        threat__status="Investigating"
    ).count()

    resolved_count = alerts.filter(
        threat__status="Resolved"
    ).count()

    # --------------------------------
    # TOP ATTACKERS
    # --------------------------------

    top_attackers = (
        alerts
        .values("threat__packet__source_ip")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # --------------------------------
    # TOP DESTINATIONS
    # --------------------------------

    top_destinations = (
        alerts
        .values("threat__packet__destination_ip")
        .annotate(total=Count("id"))
        .order_by("-total")[:10]
    )

    # --------------------------------
    # PROTOCOLS
    # --------------------------------

    protocols = (
        alerts
        .values("threat__packet__protocol")
        .annotate(total=Count("id"))
        .order_by("-total")
    )

    # --------------------------------
    # TIMELINE
    # --------------------------------

    timeline = (
        alerts
        .annotate(hour=TruncHour("created_at"))
        .values("hour")
        .annotate(total=Count("id"))
        .order_by("hour")
    )

    # =====================================
    # CHART DATA
    # =====================================

    severity_labels = [
        "Critical",
        "High",
        "Medium",
        "Low",
    ]

    severity_data = [
        critical_count,
        high_count,
        medium_count,
        low_count,
    ]

    timeline_labels = []
    timeline_data = []

    for item in timeline:

        if item["hour"]:

            timeline_labels.append(
                item["hour"].strftime("%H:%M")
            )

            timeline_data.append(
                item["total"]
            )

    # =====================================
    # CONTEXT
    # =====================================

    context = {

        "alerts": alerts[:100],

        "total_alerts": total_alerts,
        "today_alerts": today_alerts,

        "critical_count": critical_count,
        "high_count": high_count,
        "medium_count": medium_count,
        "low_count": low_count,

        "new_count": new_count,
        "investigating_count": investigating_count,
        "resolved_count": resolved_count,

        "top_attackers": top_attackers,
        "top_destinations": top_destinations,
        "protocols": protocols,

        "timeline": timeline,

        # Charts
        "severity_labels": json.dumps(severity_labels),
        "severity_data": json.dumps(severity_data),

        "timeline_labels": json.dumps(timeline_labels),
        "timeline_data": json.dumps(timeline_data),
    }

    return render(
        request,
        "dashboard/alerts.html",
        context
    )

# ==========================================================
# REPORTS
# ==========================================================

@login_required(login_url="login")
def reports(request):

    return render(
        request,
        "dashboard/reports.html",
    )

# ==========================================================
# SETTINGS
# ==========================================================

@login_required(login_url="login")
def settings(request):

    system_status, created = SystemStatus.objects.get_or_create(
        id=1
    )

    if request.method == "POST":

        system_status.packet_capture = (
            "packet_capture" in request.POST
        )

        system_status.threat_detection = (
            "threat_detection" in request.POST
        )

        system_status.machine_learning = (
            "machine_learning" in request.POST
        )

        system_status.alert_engine = (
            "email_alerts" in request.POST
        )

        system_status.save()

    context = {
        "system_status": system_status,
    }

    return render(
        request,
        "dashboard/settings.html",
        context,
    )
# ==========================================================
# PACKET PDF REPORT
# ==========================================================

@login_required(login_url="login")
def packet_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="packet_report.pdf"'
    )

    pdf = canvas.Canvas(response)

    pdf.drawString(
        50,
        800,
        "Network Packet Capture Report"
    )

    packets = PacketLog.objects.all()[:50]

    y = 760

    for packet in packets:

        pdf.drawString(
            50,
            y,
            f"{packet.captured_at} | "
            f"{packet.source_ip} -> "
            f"{packet.destination_ip} | "
            f"{packet.protocol}"
        )

        y -= 20

        if y < 50:
            pdf.showPage()
            y = 800

    pdf.save()

    return response

# ==========================================================
# THREAT PDF REPORT
# ==========================================================

@login_required(login_url="login")
def threat_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="threat_report.pdf"'
    )

    pdf = canvas.Canvas(response)

    pdf.drawString(
        50,
        800,
        "Threat Detection Report"
    )

    threats = Threat.objects.all()[:50]

    y = 760

    for threat in threats:

        pdf.drawString(
            50,
            y,
            f"{threat.detected_at} | "
            f"{threat.threat_name} | "
            f"{threat.severity}"
        )

        y -= 20

        if y < 50:
            pdf.showPage()
            y = 800

    pdf.save()

    return response

# ==========================================================
# ALERT PDF REPORT
# ==========================================================

@login_required(login_url="login")
def alert_report_pdf(request):

    response = HttpResponse(
        content_type="application/pdf"
    )

    response["Content-Disposition"] = (
        'attachment; filename="alert_report.pdf"'
    )

    pdf = canvas.Canvas(response)

    pdf.drawString(
        50,
        800,
        "Security Alert Report"
    )

    alerts = Alert.objects.all()[:50]

    y = 760

    for alert in alerts:

        pdf.drawString(
            50,
            y,
            f"{alert.created_at} - {alert.message}"
        )

        y -= 20

        if y < 50:
            pdf.showPage()
            y = 800

    pdf.save()

    return response

# ==========================================================
# THREAT EXCEL REPORT
# ==========================================================

@login_required(login_url="login")
def threat_report_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Threats"


    sheet.append([
        "Time",
        "Threat",
        "Severity",
        "Description"
    ])


    threats = Threat.objects.all()


    for threat in threats:

        sheet.append([
            threat.detected_at,
            threat.threat_name,
            threat.severity,
            threat.description
        ])


    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


    response["Content-Disposition"] = (
        'attachment; filename="threat_report.xlsx"'
    )


    workbook.save(response)


    return response

# ==========================================================
# ALERT EXCEL REPORT
# ==========================================================

@login_required(login_url="login")
def alert_report_excel(request):

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Alerts"


    sheet.append([
        "Time",
        "Threat",
        "Severity",
        "Message"
    ])


    alerts = (
        Alert.objects
        .select_related("threat")
    )


    for alert in alerts:

        sheet.append([

            alert.created_at,

            alert.threat.threat_name
            if alert.threat
            else "Unknown",

            alert.threat.severity
            if alert.threat
            else "Unknown",

            alert.message

        ])


    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


    response["Content-Disposition"] = (
        'attachment; filename="alert_report.xlsx"'
    )


    workbook.save(response)


    return response

from django.http import HttpResponse
import openpyxl

def packet_report_excel(request):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Packet Report"

    sheet.append([
        "Source IP",
        "Destination IP",
        "Protocol",
        "Time"
    ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="packet_report.xlsx"'
    )

    workbook.save(response)

    return response


from django.http import JsonResponse
from monitoring.models import PacketLog, Threat, Alert


@login_required(login_url="login")
def database_status(request):

    return JsonResponse({
        "packets": PacketLog.objects.count(),
        "threats": Threat.objects.count(),
        "alerts": Alert.objects.count(),
    })